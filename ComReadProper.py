import serial
import openpyxl

# Function to extract values from the input string
def extract_values(input_string):
    values = {}
    # Assuming the input string format remains constant
    parts = input_string.split()
    values['SERN'] = parts[0].lstrip("b'\\rSN:")
    values['DATE'] = parts[1]
    values['TIME'] = parts[2]
    values['MODE'] = parts[3]
    values['DURA'] = parts[4].rstrip("sec")
    values['ANGL'] = parts[5].rstrip("degrees")
    values['SCRW'] = parts[6].lstrip("Screw:")
    values['ERRO'] = parts[7].split('/')[0].lstrip("Errors:")
    values['PROG'] = parts[8].lstrip("Prog:")
    values['TORQ'] = parts[9].lstrip("Torque:").rstrip("Nm")
    values['SEQE'] = parts[10].lstrip("Seq:").rstrip("\n")

    return values

#COM COM port configuration
#COM comport = 'COM3'
#COM baudrate = 57600
#COM timeout = 1

# Excel file configuration
excel_file_path = 'C:/Users/Christian/Desktop/MECCATRONICA/Tesi/codetests/comread/output_data.xlsx'
sheet_name = 'Data'

#COM Open COM port
#COM ser = serial.Serial(port=comport, baudrate=baudrate, bytesize=8, parity='N', stopbits=1, timeout=timeout)

#COM Read data from COM port
#COM input_string = ser.readline()


# Specify the path to your text file
file_path = 'C:/Users/Christian/Desktop/MECCATRONICA/Tesi/codetests/comread/strings.txt'

    
        
with open(file_path, 'r') as file:
    try:
        # Load existing workbook or create a new one
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select or create the sheet
    sheet = workbook.get_sheet_by_name(sheet_name) if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # Read each line from the file
    for row_num, line in enumerate(file, start=sheet.max_row + 1):
        # Strip removes leading and trailing whitespaces
        input_string = line.strip()
        print(input_string)
        # Extract values
        data_values = extract_values(input_string)
        print(data_values)
        print('Serial:' + data_values['SERN'])
        print('Screw:' + data_values['SCRW'])
        print('Errors:' + data_values['ERRO'])
        
        # Write keys in the first row
        for col_num, key in enumerate(data_values.keys(), start=1):
            sheet.cell(row=1, column=col_num, value=key)

        # Write values in subsequent rows
        for col_num, value in enumerate(data_values.values(), start=1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Save the workbook
    workbook.save(excel_file_path)

# Close the COM port
# ser.close()
