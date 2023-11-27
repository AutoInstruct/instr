import openpyxl

#This function takes data strings from electric screwdrivers and appends them to the first empty row of an excel file
#It makes use of the extract_values function to label the first row with the data keys and put each value ina different column
#The arguments are the string to append, the path to the excel file and the name of the sheet in the file

def stringtoexcel(datastring, excelpath, sheetname):
    # Excel file configuration
    excel_file_path = excelpath #'C:/Users/Christian/Desktop/MECCATRONICA/Tesi/codetests/comread/output_data.xlsx'
    sheet_name = sheetname
    
    try:
        # Load existing workbook or create a new one
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select or create the sheet
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # Read the input string
    # Strip removes leading and trailing whitespaces
    input_string = datastring
    # Extract values
    data_values = extract_values(input_string)
    print(data_values)
    # Write keys in the first row
    for col_num, key in enumerate(data_values.keys(), start=1):
        sheet.cell(row=1, column=col_num, value=key)

    # Write values in subsequent rows
    newrow = sheet.max_row + 1
    for col_num, value in enumerate(data_values.values(), start=1):
        sheet.cell(row=newrow, column=col_num, value=value)

    # Save the workbook
    workbook.save(excel_file_path)